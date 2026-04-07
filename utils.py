import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
from matplotlib.ticker import MultipleLocator


class Exporter:
    #initialize
    def __init__(self, results, algorithms, theme, colors, algorithm_descriptions):
        self.results = results
        self.algorithms = algorithms
        self.theme = theme
        self.colors = colors
        self.algorithm_descriptions = algorithm_descriptions
        
    # gantt chart
    def save_single_gantt_chart(self, algo, data, path):
        fig, ax = plt.subplots(figsize=(11, 2.4))
        ax.set_title(f'Gantt Chart - {algo}', fontsize=12, fontweight='bold')
        ax.set_xlabel('Time Units')
        ax.set_yticks([5])
        ax.set_yticklabels(['CPU'])
        ax.grid(True, alpha=0.25, axis='x')
        for entry in data['gantt']:
            duration = entry.end_time - entry.start_time
            if entry.pid == -1:
                ax.broken_barh([(entry.start_time, duration)], (0, 10), facecolors=self.theme['idle'], hatch='//', edgecolor='#94a3b8')
                ax.text(entry.start_time + duration / 2, 5, 'IDLE', ha='center', va='center', fontsize=8)
            else:
                color = self.colors[(entry.pid - 1) % len(self.colors)]
                ax.broken_barh([(entry.start_time, duration)], (0, 10), facecolors=color, edgecolor='white', linewidth=1.1)
                ax.text(entry.start_time + duration / 2, 5, f'P{entry.pid}', ha='center', va='center', color='white', fontweight='bold')
        ax.set_xlim(0, max(1, data['total_time'] + 1))
        ax.xaxis.set_major_locator(MultipleLocator(1))
        fig.tight_layout()
        fig.savefig(path, dpi=160, bbox_inches='tight')
        plt.close(fig)
    
    #excel
    def export_to_excel(self, filepath, fig_comp):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'

        header_fill = PatternFill('solid', fgColor='2563EB')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF')
        summary_fill = PatternFill('solid', fgColor='2563EB')
        summary_font = Font(name='Calibri', bold=True, color='FFFFFF')
        normal_font = Font(name='Calibri', color='000000')
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        thin = Side(style='thin', color='D1D5DB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws['B2'] = 'CPU Scheduling Simulator Results'
        ws['B2'].font = Font(name='Calibri', size=16, bold=True)
        ws['B3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['B4'] = 'Algorithms compared: ' + ', '.join(self.algorithms)

        summary_headers = ['Algorithm', 'Avg Turnaround', 'Avg Waiting', 'Avg Response', 'CPU Utilization %', 'Throughput', 'Busy Time', 'Idle Time', 'Active Span']
        start_row = 6
        for c, h in enumerate(summary_headers, start=2):
            cell = ws.cell(row=start_row, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        row = start_row + 1
        for algo in self.algorithms:
            data = self.results[algo]
            vals = [algo, data['atat'], data['awt'], data['art'], data['util'], data['thr'], data['busy_time'], data['idle_time'], data['active_span']]
            for c, v in enumerate(vals, start=2):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = normal_font
                cell.alignment = center if c > 2 else left
                cell.border = border
            row += 1

        for col in range(2, 11):
            ws.column_dimensions[get_column_letter(col)].width = 18

        for algo in self.algorithms:
            sh = wb.create_sheet(title=algo[:31])
            sh['B2'] = f'{algo} Results'
            sh['B2'].font = Font(name='Calibri', size=15, bold=True)
            sh['B3'] = self.algorithm_descriptions[algo]
            sh['B3'].alignment = Alignment(wrap_text=False)
            sh.row_dimensions[3].height = 38

            headers = ['PID', 'Arrival', 'Burst', 'Priority', 'Start', 'Completion', 'TAT', 'WT', 'RT']
            r0 = 5
            for c, h in enumerate(headers, start=2):
                cell = sh.cell(row=r0, column=c, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            rr = r0 + 1
            for p in self.results[algo]['procs']:
                vals = [p.pid, p.arrival_time, p.burst_time, p.priority, p.start_time, p.completion_time, p.turnaround_time, p.waiting_time, p.response_time]
                for c, v in enumerate(vals, start=2):
                    cell = sh.cell(row=rr, column=c, value=v)
                    cell.font = normal_font
                    cell.alignment = center
                    cell.border = border
                rr += 1

            avg_vals = ['AVG', '', '', '', '', '', self.results[algo]['atat'], self.results[algo]['awt'], self.results[algo]['art']]
            for c, v in enumerate(avg_vals, start=2):
                cell = sh.cell(row=rr, column=c, value=v)
                cell.font = summary_font
                cell.fill = summary_fill
                cell.alignment = center
                cell.border = border
                
                
            img_path = os.path.join(os.path.dirname(filepath), f'{algo.replace(" ", "_").replace("/", "_")}_gantt.png')
            self.save_single_gantt_chart(algo, self.results[algo], img_path)
            if os.path.exists(img_path):
                img = XLImage(img_path)
                img.width = 900
                img.height = 180
                sh.add_image(img, 'B17')

            for col in range(2, 11):
                sh.column_dimensions[get_column_letter(col)].width = 14

        comp_img = os.path.join(os.path.dirname(filepath), 'comparison_chart.png')
        fig_comp.savefig(comp_img, dpi=160, bbox_inches='tight')
        if os.path.exists(comp_img):
            img = XLImage(comp_img)
            img.width = 900
            img.height = 520
            ws.add_image(img, 'B15')

        wb.save(filepath)
    
    # pdf
    def export_to_pdf(self, filepath, fig_comp):
        with PdfPages(filepath) as pdf:
            # Page 1: Summary
            fig = plt.figure(figsize=(11.69, 8.27), constrained_layout=True)
            fig.patch.set_facecolor('white')
            ax = fig.add_subplot(111)
            ax.axis('off')
            
            ax.text(0.5, 0.98, 'CPU Scheduling Simulator Results', va='top', ha='center', 
                    fontsize=22, weight='bold', family='sans-serif')
            ax.text(0.5, 0.94, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 
                    va='top', ha='center', fontsize=11, color='#666666')
            
            columns = ('Algorithm', 'Avg TAT', 'Avg WT', 'Avg RT', 'Util %', 'Throughput')
            table_data = []
            for algo in self.algorithms:
                d = self.results[algo]
                table_data.append([
                    algo, f"{d['atat']:.2f}", f"{d['awt']:.2f}", f"{d['art']:.2f}", 
                    f"{d['util']:.1f}%", f"{d['thr']:.3f}"
                ])

            main_table = ax.table(
                cellText=table_data, colLabels=columns,
                cellLoc='center', loc='center', bbox=[0.05, 0.52, 0.90, 0.35]
            )
            main_table.auto_set_font_size(False)
            main_table.set_fontsize(10)
            
            for (row, col), cell in main_table.get_celld().items():
                cell.set_edgecolor('#D1D5DB')
                cell.set_linewidth(0.6)
                if row == 0:
                    cell.set_text_props(weight='bold', color='white')
                    cell.set_facecolor('#1E40AF')
                else:
                    cell.set_facecolor('#F9FAFB' if row % 2 == 0 else '#FFFFFF')

            pdf.savefig(fig)
            plt.close(fig)

            # Page 2: Comparison
            if fig_comp:
                fig_comp.set_size_inches(11.69, 8.27)
                pdf.savefig(fig_comp, bbox_inches='tight')

            # Pages 3+: Detailed analysis
            for algo in self.algorithms:
                d = self.results[algo]
                
                fig, ax = plt.subplots(figsize=(11.69, 8.27))
                ax.axis('off')
                ax.text(0.5, 0.98, f"Detailed Analysis: {algo}", ha='center', fontsize=18, weight='bold')
                
                proc_columns = ['PID', 'Arrival', 'Burst', 'Prio', 'Start', 'Comp', 'TAT', 'WT', 'RT']
                proc_data = [[
                    p.pid, f"{p.arrival_time:.1f}", f"{p.burst_time:.0f}", p.priority,
                    f"{p.start_time:.1f}", f"{p.completion_time:.1f}",
                    f"{p.turnaround_time:.1f}", f"{p.waiting_time:.1f}", f"{p.response_time:.1f}"
                ] for p in d['procs']]
                
                proc_data.append([
                    'AVG', '', '', '', '', '',
                    f"{d['atat']:.2f}", f"{d['awt']:.2f}", f"{d['art']:.2f}"])

                p_table = ax.table(cellText=proc_data, colLabels=proc_columns, cellLoc='center', loc='center', bbox=[0.05, 0.52, 0.90, 0.35])
                
                last_table_row = len(proc_data)
                
                for (row, col), cell in p_table.get_celld().items():
                    cell.set_edgecolor('#D1D5DB')
                    cell.set_linewidth(0.6)
                    if row == 0:
                        cell.set_text_props(weight='bold', color='white')
                        cell.set_facecolor('#1E40AF')
                    elif row == last_table_row:
                            cell.set_text_props(weight='bold', color='#1E3A8A')
                            cell.set_facecolor('#DBEAFE')
                    else:
                        cell.set_facecolor('#F9FAFB' if row % 2 == 0 else '#FFFFFF')
                
                p_table.auto_set_font_size(False)
                p_table.set_fontsize(10)
                
                pdf.savefig(fig)
                plt.close(fig)

                # Gantt page
                temp_gantt = f"temp_{algo}.png"
                self.save_single_gantt_chart(algo, d, temp_gantt)
                
                if os.path.exists(temp_gantt):
                    fig_g, ax_g = plt.subplots(figsize=(11.69, 4))
                    img = plt.imread(temp_gantt)
                    ax_g.imshow(img)
                    ax_g.set_title(f"{algo} Execution Timeline", fontsize=14, pad=10)
                    ax_g.axis('off')
                    pdf.savefig(fig_g, bbox_inches='tight', dpi=300)
                    plt.close(fig_g)
                    os.remove(temp_gantt)

# import csv instead of hard key each time
def import_csv(filepath):
    """Import processes from CSV file"""
    import csv
    from models import Process
    
    processes = []
    with open(filepath, 'r', newline='') as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, start=1):
            arrival = float(row['Arrival Time'])
            burst = int(float(row['Burst Time']))
            priority = int(row['Priority'])
            processes.append(Process(i, arrival, burst, priority))
    return processes


def export_csv(filepath, processes):
    """Export processes to CSV file"""
    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['PID', 'Arrival Time', 'Burst Time', 'Priority'])
        for p in processes:
            writer.writerow([p.pid, p.arrival_time, p.burst_time, p.priority])